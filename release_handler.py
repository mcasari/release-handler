import yaml
import os
import subprocess
import logging
import sys
import xml.etree.ElementTree as ET
import re
import click
import platform
from lxml import etree as lxmlET
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from git import Repo

# Configure logging
logging.basicConfig(filename='release-handler.log', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')
def _has_special_characters(s):
    return bool(re.search(r'[^a-zA-Z0-9]', s))
    
def _refresh_git_tags(repo_path):
    """
    Deletes all local tags and fetches remote tags in a Git repository.

    :param repo_path: Path to the local Git repository.
    """
    if not os.path.isdir(repo_path):
        raise ValueError(f"Path '{repo_path}' is not a directory")

    # Change to the git repo directory
    original_dir = os.getcwd()
    os.chdir(repo_path)

    try:
        # Verify it's a git repository
        subprocess.run(["git", "rev-parse", "--is-inside-work-tree"], check=True, stdout=subprocess.DEVNULL)

        # Get all local tags
        result = subprocess.run(["git", "tag"], check=True, stdout=subprocess.PIPE, text=True)
        tags = result.stdout.strip().split("\n")

        # Delete local tags
        
        if tags and tags[0] != '':
            for tag in tags:
                if _has_special_characters(tag):
                    print(f"remove tag {tag}")
                    tags.remove(tag)                    
            # Run subprocess with UTF-8 encoding
            subprocess.run(["git", "tag", "-d"] + tags, check=True, text=True, encoding="utf-8")
            print(f"Deleted local tags {tags}")
        else:
            print("No local tags to delete.")

        # Fetch remote tags
        subprocess.run(["git", "fetch", "--tags"], check=True)
        print("Fetched remote tags.")

    except subprocess.CalledProcessError as e:
        print(f"An error occurred: {e}")
    finally:
        os.chdir(original_dir)
                    
def _run_git_command(repo_path, args):
    try:
        result = subprocess.check_output(["git", "-C", repo_path] + args, stderr=subprocess.DEVNULL)
        return result.decode("utf-8").strip()
    except subprocess.CalledProcessError:
        return ""

def _get_git_info(repo_path):
    if not os.path.exists(os.path.join(repo_path, ".git")):
        return {"Path": repo_path, "Error": "Not a git repository"}

    remote_url = _run_git_command(repo_path, ["remote", "get-url", "origin"])
    last_commit = _run_git_command(repo_path, ["rev-parse", "HEAD"])
    commit_msg = _run_git_command(repo_path, ["log", "-1", "--pretty=%s"])
    commit_date = _run_git_command(repo_path, ["log", "-1", "--date=iso", "--pretty=%cd"])
    tags = _run_git_command(repo_path, ["tag", "--points-at", last_commit]).splitlines()
    tags = ", ".join(tags) if tags else "None"

    # Get all branches that point to the same commit as one of the tags (if any)
    branch_map = {}
    if tags != "None":
        tag_list = tags.split(", ")
        for tag in tag_list:
            tag_commit = _run_git_command(repo_path, ["rev-list", "-n", "1", tag])
            branches = _run_git_command(repo_path, ["branch", "--contains", tag_commit]).replace("*", "").splitlines()
            branch_map[tag] = [b.strip() for b in branches]
    else:
        branch_map = {}

    return {
        "Path": repo_path,
        "Remote": remote_url,
        "Last Commit": last_commit,
        "Commit Message": commit_msg,
        "Commit Date": commit_date,
        "Tags": tags,
        "Branches with Same Commit as Tag": str(branch_map)
    }
         
        
def _is_tag_committed(tag_name, repo_path):
    """
    Check if a given Git tag exists in the repository.

    Args:
        tag_name (str): The name of the tag to check.
        repo_path (str): Path to the Git repository (default is current directory).

    Returns:
        bool: True if the tag exists (i.e., is committed), False otherwise.
    """
    try:
        result = subprocess.run(
            ['git', 'rev-parse', '--verify', f'refs/tags/{tag_name}'],
            cwd=repo_path,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )
        return result.returncode == 0
    except Exception as e:
        print(f"Error checking tag: {e}")
        return False        
 
def _list_git_changes(project_path):
    """
    Lists all tracked changes (modified, added, deleted) in a Git repository.
    
    :param project_path: Path to the Git project.
    :return: A dictionary containing lists of modified, added, and deleted files.
    """
    if not os.path.isdir(project_path):
        raise ValueError("Invalid project path")
    
    git_command = ["git", "status", "--porcelain"]
    try:
        result = subprocess.run(git_command, cwd=project_path, capture_output=True, text=True, check=True)
        changes = result.stdout.strip().split('\n')
        
        modified = []
        added = []
        deleted = []
        
        for change in changes:
            if change:
                status, file = change[:2].strip(), change[3:].strip()
                if status in ('M', 'MM'):
                    modified.append(file)
                elif status in ('A', 'AM'):
                    added.append(file)
                elif status in ('D', 'DM'):
                    deleted.append(file)
        
        return {"modified": modified, "added": added, "deleted": deleted}
    
    except subprocess.CalledProcessError as e:
        raise RuntimeError(f"Error executing git command: {e}")
        
def _is_tag_pushed(project_path, tag_name) -> bool:
    """
    Checks if a given tag is already pushed to the remote repository.
    
    :param project_path: Path to the local git repository.
    :param tag_name: Name of the tag to check.
    :return: True if the tag is pushed, False otherwise.
    """
    try:
        # Get the list of tags from the remote
        result = subprocess.run(
            ["git", "ls-remote", "--tags", "origin"],
            cwd=project_path,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=True
        )
        
        # Check if the tag exists in the remote
        return any(f"refs/tags/{tag_name}" in line for line in result.stdout.splitlines())
    except subprocess.CalledProcessError as e:
        print(f"Error checking remote tags: {e.stderr}")
        return False
        
def _compile_maven_project(project_path, maven_home, settings_file, config) -> bool:
    """
    Compiles a Maven project while skipping tests.
    
    :param project_path: Path to the Maven project.
    :param maven_home: Path to the Maven home directory.
    :param settings_file: Path to the Maven settings file.
    :return: True if compilation succeeds, False otherwise.
    """
    is_windows = platform.system() == "Windows"
    mvn_executable = os.path.join(maven_home, "bin", "mvn.cmd" if is_windows else "mvn")
    command = [mvn_executable, "clean", "compile", "--settings", settings_file]
    maven_compile_options = config.get("maven_compile_options", [])
    install_index = command.index("compile")
    command = command[:install_index + 1] + maven_compile_options + command[install_index + 1:]
    
    try:
        result = subprocess.run(command, cwd=project_path, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        if result.returncode == 0:
            return True
        else:
            logging.info("Maven build failed:", result.stderr)
            print("Maven build failed:", result.stderr)
            return False
    except Exception as e:
        logging.error(f"Error running Maven: {e}")
        print(f"Error running Maven: {e}")
        return False
        
def _compile_angular_project(project_path, config) -> bool:
    """
    Checks if an Angular project compiles correctly.
    :param project_path: Path to the Angular project.
    :return: True if the project compiles successfully, False otherwise.
    """
    if not os.path.isdir(project_path):
        logging.error("Invalid project path.")
        print("Invalid project path.")
        return False
    
    try:
        # Run the Angular build command
        npm_command = os.path.join(config["nodejs_home"], "ng")
        is_windows = platform.system() == "Windows"
        if is_windows:
            npm_command = os.path.join(config["nodejs_home"], "ng.cmd")
        nodejs_compile_options = config.get("nodejs_compile_options", [])
        npm_command_arr = [npm_command, "build"] + nodejs_compile_options;
        result = subprocess.run(
            npm_command_arr, cwd=project_path, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True
        )
        if result.returncode == 0:
            return True
        else:
            logging.info("Build failed:", result.stderr)
            print("Build failed:", result.stderr)
            return False
    except FileNotFoundError as e:
        logging.error(f"Error: {e}")
        print(f"Error: {e}")
        return False

# Example usage:
# print(check_angular_compile("/path/to/angular/project"))

def _compile_ant_project(project_path, config) -> bool:
    """
    Compiles an Ant project.
    
    :param project_path: Path to the Ant project.
    :param ant_home: Path to the Ant home directory.
    :param ant_target: The Ant target to execute.
    :return: True if compilation is successful, False otherwise.
    """
    ant_executable = os.path.join(config["ant_home"], 'bin', 'ant')
    is_windows = platform.system() == "Windows"
    if is_windows:
        ant_executable = os.path.join(config["ant_home"], 'bin', 'ant.bat')
    command = [ant_executable, config["ant_target"]] + config["ant_compile_options"]
    
    try:
        result = subprocess.run(command, cwd=project_path, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        logging.info(result.stderr)
        print(result.stderr)
        if "failed" in result.stderr:
            raise ValueError("Unexpected error during build: " + result.stderr)       
        return result.returncode == 0
    except Exception as e:
        logging.error(f"Error executing Ant: {e}")
        print(f"Error executing Ant: {e}")
        return False

# Example usage:
# success = compile_ant_project("/path/to/project", "/path/to/ant", "build")
# print("Compilation Successful" if success else "Compilation Failed")

def _is_last_commit_pushed(project_path):
    try:
        # Get the latest commit hash
        commit_hash = subprocess.check_output(
            ["git", "rev-parse", "HEAD"], cwd=project_path
        ).strip().decode("utf-8")

        # Check if the commit is in the remote
        remote_branches = subprocess.check_output(
            ["git", "branch", "-r", "--contains", commit_hash], cwd=project_path
        ).strip().decode("utf-8")

        # If the commit exists in remote branches, it's pushed
        return bool(remote_branches)
    except subprocess.CalledProcessError:
        return False              
                                        
def _resolve_placeholders(data, context=None):
    """ Recursively resolves placeholders in a dictionary using string formatting """
    if context is None:
        context = data  # Initial context is the entire data
    
    if isinstance(data, dict):
        return {key: _resolve_placeholders(value, context) for key, value in data.items()}
    elif isinstance(data, list):
        return [_resolve_placeholders(item, context) for item in data]
    elif isinstance(data, str):
        try:
            return data.format(**context)  # Use Python string formatting
        except KeyError:
            return data  # Return as-is if placeholders are unresolved
    else:
        return data  # Return non-string values unchanged
                    
def _update_all_pom_properties(project, config):
    for root, _, files in os.walk(project["project_path"]):
        for file in files:
            if file == "pom.xml":
                file_path = os.path.join(root, file)
                _update_pom_property(file_path, project, config)

def _update_pom_property(file_path, project, config):
    # Parse XML while preserving comments
    parser = lxmlET.XMLParser(remove_comments=False)
    tree = lxmlET.parse(file_path, parser)
    root = tree.getroot()
    maven_namespace = config["maven_namespace"]
    
    # Define the XML namespace
    ns = {'maven': maven_namespace}
    ET.register_namespace('', ns['maven'])
    
    # Locate the <properties> section
    modified = False
    properties = root.find(".//maven:properties", ns)
    if properties is not None:     
        if project["type"] == "Maven":
            for configProperties in project["properties"]:                 
                property_element = properties.find(f"maven:{configProperties['property_name']}", ns)
                if property_element is not None:
                    property_element.text = configProperties['property_value']
                    logging.info(f"Updated {configProperties['property_name']} to {configProperties['property_value']}")
                    print(f"Updated {configProperties['property_name']} to {configProperties['property_value']}")
                    modified = True                
                else:
                    logging.info(f"Property {configProperties['property_name']} not found in pom.xml.")
                    print(f"Property {configProperties['property_name']} not found in pom.xml.")
    else:
        logging.info("No <properties> section found in pom.xml.")
        print("No <properties> section found in pom.xml.")
    if modified:
        tree.write(file_path, xml_declaration=True, encoding='utf-8')
        
def _update_maven_versions(project_path, dependencies, version, parent_version, maven_namespace):
    for root, _, files in os.walk(project_path):
        if 'pom.xml' in files:
            pom_path = os.path.join(root, 'pom.xml')
            # Parse XML while preserving comments
            parser = lxmlET.XMLParser(remove_comments=False)
            tree = lxmlET.parse(pom_path, parser)
            
            root_element = tree.getroot()
            
            # Define the XML namespaces
            # Define the XML namespace
            namespaces = {'m': maven_namespace}
            ET.register_namespace('', namespaces['m'])
            
            modified = False

            # Update the version of the project itself
            for project_version in root_element.findall("./m:version", namespaces):
                if project_version is not None:
                    project_version.text = version
                    modified = True
                    logging.info(f"Project version updated to {version}")
                    print(f"Project version updated to {version}")

            # Update the parent version
            for parent_tag in root_element.findall("./m:parent", namespaces):
                if parent_tag is not None:
                    for parent_vers in parent_tag.findall("./m:version", namespaces):                      
                        if parent_vers is not None:
                            parent_vers.text = parent_version
                            modified = True
                            logging.info(f"Parent version updated to {version}")
                            print(f"Parent version updated to {version}")
            
            
            for dependency in root_element.findall(".//m:dependency", namespaces):
                artifact_id_elem = dependency.find("m:artifactId", namespaces)
                version_elem = dependency.find("m:version", namespaces)
                
                if artifact_id_elem is not None and version_elem is not None:
                    for dependency in dependencies:
                        if artifact_id_elem.text in dependency["dependency_name"]:
                            version_elem.text = dependency["dependency_version"]
                            modified = True
                            logging.info(f"Dependency {dependency["dependency_name"]}  updated to {dependency["dependency_version"]}")
                            print(f"Dependency {dependency["dependency_name"]}  updated to {dependency["dependency_version"]}")            
            if modified:
                tree.write(pom_path, xml_declaration=True, encoding='utf-8')

def _update_maven_versions_from_yaml(project, config):    
    maven_namespace = config["maven_namespace"]
    project_path = project.get("project_path")
    parent_version = project.get("parent_version")
    dependencies = project.get("dependencies", [])
    version = project.get("version")
    if not project_path or not dependencies or not version or not maven_namespace:
        raise ValueError("YAML file must contain 'maven_namespace', 'project_path', 'parent_version', 'dependencies' and 'version' fields.")
    
    _update_maven_versions(project_path, dependencies, version, parent_version, maven_namespace)

                                       
                    
def _find_file(base_path, filename):
    for root, _, files in os.walk(base_path):
        if filename in files:
            return os.path.join(root, filename)  
    return None
    
def _execute_command(command, cwd):
    """Executes a shell command in a given directory."""
    try:
        subprocess.run(command, cwd=cwd, check=True, shell=True)
        logging.info(f"Executed: {' '.join(command)} in {cwd}")
        print(f"Executed: {' '.join(command)} in {cwd}")
    except subprocess.CalledProcessError as e:
        logging.error(f"Command failed: {e}")
        print(f"Command failed: {e}")
        raise

def _update_ant_version(path, version, version_file):
    """Updates the Ant project version."""
    version_file_path = _find_file(path, version_file)
    with open(version_file_path, 'r') as file:
        content = file.read()
    content = re.sub(r'(?m)^\s*version\s*=\s*.*$', f'version = {version}', content)
    with open(version_file_path, 'w') as file:
        file.write(content)
    print(f"Updated Ant version in {version_file_path} to {version}")
    logging.info(f"Updated Ant version in {version_file_path} to {version}")

def _update_angular_version(path, version, version_file):
    """Updates the Angular project version."""
    version_file_path = _find_file(path, version_file)
    with open(version_file_path, 'r') as file:
        content = file.read()
    content = re.sub(r'"version"\s*:\s*".*?"\s*,', f'"version": "{version}",', content)
    with open(version_file_path, 'w') as file:
        file.write(content)
    print(f"Updated Angular version in {version_file_path} to {version}")
    logging.info(f"Updated Angular version in {version_file_path} to {version}")
    
def checkout_and_pull(project_filter = ''):
    """Performs git checkout on master and pulls latest changes."""
    with open("release_handler_config.yaml", "r") as file:
        config = yaml.safe_load(file)   
    try:
        for project in config["projects"]:
            if project_filter != '' and project_filter != project['name']:
                continue  
            if 'skip' in project and project['skip']:
                logging.info(f"Project {project['name']} is configured to be skipped")
                print(f"Project {project['name']} is configured to be skipped")
                continue
            if click.confirm(f"Check out and pull project {project['name']}?", default=True):
                branch = project["git_branch"]
                _execute_command(["git", "checkout", branch], project['project_path'])
                _execute_command(["git", "pull"], project['project_path'])
                logging.info(f"Project {project['name']} checked out and pulled")  
                print(f"Project {project['name']} checked out and pulled")
    except Exception as e:
        logging.error(f"An error occurred: {e}")  
        print(f"An error occurred: {e}")

def update_versions(project_filter = ''):
    """Reads the YAML file and processes each project."""
    with open("release_handler_config.yaml", "r") as file:
        config = yaml.safe_load(file)
    
    try:
        for project in config["projects"]:
            if project_filter != '' and project_filter != project['name']:
                continue  
            if 'skip' in project and project['skip']:
                logging.info(f"Project {project['name']} is configured to be skipped")
                print(f"Project {project['name']} is configured to be skipped")
                continue
            _execute_command(["git", "checkout", project["git_branch"]], project['project_path'])
            if click.confirm(f"Update version for project {project['name']}?", default=True):
                if project["type"] == "Maven":
                    _update_all_pom_properties(project, config)
                    _update_maven_versions_from_yaml(project, config)
                elif project["type"] == "Ant":
                    _update_ant_version(project["project_path"], project["version"], project["version_file"])
                elif project["type"] == "Angular":
                    _update_angular_version(project["project_path"], project["version"], project["version_file"])
    except Exception as e:
        logging.error(f"An error occurred: {e}")  
        print(f"An error occurred: {e}") 
    
def create_tags(project_filter = ''):
    """Tags each project with the appropriate tag name."""
    try:
        with open("release_handler_config.yaml", "r") as file:
            config = yaml.safe_load(file)
            resolved_config = _resolve_placeholders(config)
        
        for project in resolved_config["projects"]:
            if project_filter != '' and project_filter != project['name']:
                continue  
            if 'skip' in project and project['skip']:
                logging.info(f"Project {project['name']} is configured to be skipped")
                print(f"Project {project['name']} is configured to be skipped")
                continue
            _execute_command(["git", "checkout", project["git_branch"]], project['project_path'])
            _refresh_git_tags(project["project_path"])
            tag = project["tag"]
            if click.confirm(f"Create tag {tag} for project {project['name']}?", default=True):
                if _is_tag_committed(tag, project["project_path"]):
                    logging.info(f"Tag {tag} of project {project['name']} already commited")
                    print(f"Tag {tag} of project {project['name']} already commited")
                    continue               
                _execute_command(["git", "tag", tag], project["project_path"])
                logging.info(f"Tagged {project['name']} with {tag}")
                print(f"Tagged {project['name']} with {tag}")
    except Exception as ex:
        logging.error(f"An error occurred: {ex}")  
        print(f"An error occurred: {ex}")
        
        
def push_tags(project_filter = ''):
    try:
        with open("release_handler_config.yaml", "r") as file:
            config = yaml.safe_load(file)
            resolved_config = _resolve_placeholders(config)
        
        for project in resolved_config["projects"]:
            if project_filter != '' and project_filter != project['name']:
                continue  
            if 'skip' in project and project['skip']:
                logging.info(f"Project {project['name']} is configured to be skipped")
                print(f"Project {project['name']} is configured to be skipped")
                continue
            tag = project["tag"]
            if click.confirm(f"Push tag {tag} for project {project['name']}?", default=True):
                try:
                    _execute_command(["git", "checkout", project["git_branch"]], project['project_path'])
                    _refresh_git_tags(project["project_path"])
                    if not _is_tag_committed(tag, project["project_path"]):
                        _execute_command(["git", "tag", tag], project["project_path"])                            
                    if _is_tag_pushed(project["project_path"], tag):
                        logging.info(f"The {tag} for project {project['name']} is already pushed")
                        print(f"The {tag} for project {project['name']} is already pushed")
                        continue
                    _execute_command(["git", "push", "origin", "tag", tag], project["project_path"])
                except Exception as e:
                    pass
                logging.info(f"Pushed tag {tag} for project {project['name']}")
                print(f"Pushed tag {tag} for project {project['name']}")
    except Exception as ex:
        logging.error(f"An error occurred: {ex}")  
        print(f"An error occurred: {ex}")
                
def delete_tags(project_filter = ''):
    """Delete tag of each project with the appropriate tag name."""
    try:
        with open("release_handler_config.yaml", "r") as file:
            config = yaml.safe_load(file)
            resolved_config = _resolve_placeholders(config)
        
        for project in resolved_config["projects"]:
            if project_filter != '' and project_filter != project['name']:
                continue  
            if 'skip' in project and project['skip']:
                logging.info(f"Project {project['name']} is configured to be skipped")
                print(f"Project {project['name']} is configured to be skipped")
                continue
            tag = project["tag"]
            if click.confirm(f"Delete tag {tag} for project {project['name']}?", default=True):
                _execute_command(["git", "checkout", project["git_branch"]], project['project_path'])
                _refresh_git_tags(project["project_path"])
                if not _is_tag_committed(tag, project["project_path"]):
                    logging.info(f"There is no tag {tag} for project {project['name']}")
                    print(f"There is no tag {tag} for project {project['name']}")
                    continue
                _execute_command(["git", "tag", "-d", tag], project["project_path"])
                logging.info(f"Deleted tag {tag}")
                print(f"Deleted tag {tag}")
    except Exception as e:
        logging.error(f"An error occurred: {e}")  
        print(f"An error occurred: {e}") 
        
def delete_tags_remotely(project_filter = ''):
    """Delete tag of each project with the appropriate tag name."""
    try:
        with open("release_handler_config.yaml", "r") as file:
            config = yaml.safe_load(file)
            resolved_config = _resolve_placeholders(config)
        
        remote = config["remote_git_repo"]
        print(f"remote repo {remote}")
        if not remote:
            remote = "origin"
        print(f"project_filter {project_filter}")
        for project in resolved_config["projects"]:
            if project_filter != '' and project_filter != project['name']:
                continue  
            if 'skip' in project and project['skip']:
                logging.info(f"Project {project['name']} is configured to be skipped")
                print(f"Project {project['name']} is configured to be skipped")
                continue
            tag = project["tag"]
            if click.confirm(f"Delete tag {tag} remotely for project {project['name']}?", default=True):
                _execute_command(["git", "checkout", project["git_branch"]], project['project_path'])
                if not _is_tag_pushed(project["project_path"], tag):
                    logging.info(f"The {tag} for project {project['name']} does not exist remotely")
                    print(f"The {tag} for project {project['name']} does not exist remotely")
                    continue
                _execute_command(["git", "push", "--delete", remote, tag], project["project_path"])
                logging.info(f"Deleted tag {tag} remotely")
                print(f"Deleted tag {tag} remotely")
                _refresh_git_tags(project["project_path"])
    except Exception as e:
        logging.error(f"An error occurred: {e}")  
        print(f"An error occurred: {e}") 
        
def commit(project_filter = ''):
    """Commits changes for each project with confirmation."""
    try:
        with open("release_handler_config.yaml", "r") as file:
            config = yaml.safe_load(file)
        
        for project in config["projects"]:
            if project_filter != '' and project_filter != project['name']:
                continue  
            if 'skip' in project and project['skip']:
                logging.info(f"Project {project['name']} is configured to be skipped")
                print(f"Project {project['name']} is configured to be skipped")
                continue
            _execute_command(["git", "checkout", project["git_branch"]], project['project_path'])
            changes =_list_git_changes(project["project_path"])
            logging.info(f"Changes to commit {changes}")
            print(f"Changes to commit {changes}")
            if click.confirm(f"Commit changes for project {project['name']}?", default=False):
                _execute_command(["git", "commit", "-am", f"Update project with version {project['version']}"] , project["project_path"])
                logging.info(f"Update project with version {project['version']}")
                print(f"Update project with version {project['version']}")
    except Exception as e:
        logging.error(f"An error occurred: {e}")  
        print(f"An error occurred: {e}") 
        
def remove_last_commit(project_filter = ''):
    """Resets the last commit based on reset-type."""
    try:
        with open("release_handler_config.yaml", "r") as file:
            config = yaml.safe_load(file)
        
        for project in config["projects"]:
            if project_filter != '' and project_filter != project['name']:
                continue              
            if 'skip' in project and project['skip']:
                logging.info(f"Project {project['name']} is configured to be skipped")
                print(f"Project {project['name']} is configured to be skipped")
                continue
            _execute_command(["git", "checkout", project["git_branch"]], project['project_path'])
            if click.confirm(f"Reset last commit for {project['name']}?", default=False):
                if not _is_last_commit_pushed(project["project_path"]):
                    _execute_command(["git", "reset", f"--{project['reset_type']}", "HEAD~1"] , project["project_path"])
                    logging.info(f"Resetted last commit for project {project['name']}")
                    print(f"Resetted last commit for project {project['name']}")
                else:
                    logging.info(f"Reset aborted because the last commit for project {project['name']} was already pushed")
                    print(f"Reset aborted because the last commit for project {project['name']} was already pushed")
    except Exception as e:
        logging.error(f"An error occurred: {e}")  
        print(f"An error occurred: {e}") 

def reset(project_filter = ''):
    """Resets projects based on reset-type."""
    try:
        with open("release_handler_config.yaml", "r") as file:
            config = yaml.safe_load(file)
        
        for project in config["projects"]:
            if project_filter != '' and project_filter != project['name']:
                continue  
            if 'skip' in project and project['skip']:
                logging.info(f"Project {project['name']} is configured to be skipped")
                print(f"Project {project['name']} is configured to be skipped")
                continue
            _execute_command(["git", "checkout", project["git_branch"]], project['project_path'])
            if click.confirm(f"Reset {project['name']}?", default=True):
                _execute_command(["git", "reset", f"--{project['reset_type']}"] , project["project_path"])
                logging.info(f"Resetted project {project['name']}")
                print(f"Resetted project {project['name']}")
    except Exception as e:
        logging.error(f"An error occurred: {e}")  
        print(f"An error occurred: {e}") 
               
def compile_check(project_filter = ''):
    """Resets projects based on reset-type."""
    try:
        with open("release_handler_config.yaml", "r") as file:
            config = yaml.safe_load(file)
            
        for project in config["projects"]:
            if project_filter != '' and project_filter != project['name']:
                continue        
            if 'skip' in project and project['skip']:
                logging.info(f"Project {project['name']} is configured to be skipped")
                print(f"Project {project['name']} is configured to be skipped")
                continue
            _execute_command(["git", "checkout", project["git_branch"]], project['project_path'])
            if click.confirm(f"Compile {project['name']}?", default=True):
                logging.info(f"Compiling project {project['name']} ...")
                print(f"Compiling {project['name']} ...")
                if project["type"] == "Maven":
                    if _compile_maven_project(project["project_path"], config["maven_home"], config["maven_settings"], config):
                        logging.info(f"Maven project {project['name']} compiled successfully")
                        print(f"Maven project {project['name']} compiled successfully")
                elif project["type"] == "Angular":
                    if _compile_angular_project(project["project_path"], config):
                        logging.info(f"Angular project {project['name']} compiled successfully")
                        print(f"Angular project {project['name']} compiled successfully")
                elif project["type"] == "Ant":
                    if _compile_ant_project(project["project_path"], config):
                        logging.info(f"Ant project {project['name']} compiled successfully")
                        print(f"Ant project {project['name']} compiled successfully")                                      
    except Exception as e:
        logging.error(f"An error occurred: {e}")  
        print(f"An error occurred: {e}") 
          
        
def extract_git_info_to_excel(project_filter='', output_file="git_info.xlsx"):
    try:
        with open("release_handler_config.yaml", "r") as file:
            config = yaml.safe_load(file)

        repo_paths = []
        for project in config["projects"]:
            if project_filter and project_filter != project['name']:
                continue
            if 'skip' in project and project['skip']:
                logging.info(f"Project {project['name']} is configured to be skipped")
                print(f"Project {project['name']} is configured to be skipped")
                continue

            repo_paths.append(project["project_path"])

        records = [info for path in repo_paths if (info := _get_git_info(path)) is not None]

        if not records:
            print("No data to write to Excel.")
            logging.warning("No data to write to Excel.")
            return

        df = pd.DataFrame(records)
        df.to_excel(output_file, index=False)

        # Open the file with openpyxl for formatting
        wb = load_workbook(output_file)
        ws = wb.active

        # Set header fill color (orange)
        header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2  # add some padding

        wb.save(output_file)
        logging.info(f"Excel file created with formatting: {output_file}")
        print(f"Excel file created with formatting: {output_file}")

    except Exception as e:
        logging.error(f"An error occurred: {e}")
        print(f"An error occurred: {e}")
                

        
if __name__ == "__main__":
    if len(sys.argv) > 1:
        if sys.argv[1] == "update_versions":
            if len(sys.argv) > 2 and sys.argv[2] != "":
                update_versions(sys.argv[2])
            else:
                update_versions()
        elif sys.argv[1] == "create_tags":
            if len(sys.argv) > 2 and sys.argv[2] != "":
                create_tags(sys.argv[2])
            else:
                create_tags()
        elif sys.argv[1] == "delete_tags":
            if len(sys.argv) > 2 and sys.argv[2] != "":
                delete_tags(sys.argv[2])
            else:
                delete_tags()
        elif sys.argv[1] == "delete_tags_remotely":
            if len(sys.argv) > 2 and sys.argv[2] != "":
                delete_tags_remotely(sys.argv[2])
            else:
                delete_tags_remotely()
        elif sys.argv[1] == "push_tags":
            if len(sys.argv) > 2 and sys.argv[2] != "":
                push_tags(sys.argv[2])
            else:
                push_tags()    
        elif sys.argv[1] == "commit":
            if len(sys.argv) > 2 and sys.argv[2] != "":
                commit(sys.argv[2])
            else:
                commit()                  
        elif sys.argv[1] == "remove_last_commit":
            if len(sys.argv) > 2 and sys.argv[2] != "":
                remove_last_commit(sys.argv[2])
            else:
                remove_last_commit() 
        elif sys.argv[1] == "reset":
            if len(sys.argv) > 2 and sys.argv[2] != "":
                reset(sys.argv[2])
            else:
                reset() 
        elif sys.argv[1] == "checkout_and_pull":
            if len(sys.argv) > 2 and sys.argv[2] != "":
                checkout_and_pull(sys.argv[2])
            else:
                checkout_and_pull()   
        elif sys.argv[1] == "compile_check":
            if len(sys.argv) > 2 and sys.argv[2] != "":
                compile_check(sys.argv[2])
            else:
                compile_check() 
        elif sys.argv[1] == "extract_git_info_to_excel":
            if len(sys.argv) > 2 and sys.argv[2] != "":
                extract_git_info_to_excel(sys.argv[2])
            else:
                extract_git_info_to_excel()   
        elif sys.argv[1] == "_refresh_git_tags":
            _refresh_git_tags('C:/eclipse-workspaces/csi/pbservwelfare')                             
        else:
             print("Wrong argument!")           
    else:
        raise ValueError("No arguments provided! Usage: release_handler.py <one of update_versions, create_tags, delete_tags, delete_tags, commit, remove_last_commit, reset, checkout_and_pull, compile_check>")