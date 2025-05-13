import yaml
import os
import subprocess
import logging
import sys
import xml.etree.ElementTree as ET
import re
import click
import platform
import shutil
import stat
from lxml import etree as lxmlET
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from git import Repo

# Configure logging
logging.basicConfig(filename='release-handler.log', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')
def _has_special_characters(s):
    return bool(re.search(r'[^a-zA-Z0-9]', s))
    
def _on_rm_error(func, path, exc_info):
    """Error handler for shutil.rmtree to handle read-only files."""
    os.chmod(path, stat.S_IWRITE)  # Try to make it writable
    func(path)
    
def _clone_git_repo_delete_existent(repo_url, repo_local_dir):
    """
    Clones a Git repository from the given HTTPS URL into the specified folder.

    Args:
        repo_url (str): The HTTPS URL of the Git repository.
        base_dir (str): The path to the folder where the repo should be cloned.

    Raises:
        ValueError: If the repo_url is not a valid HTTPS Git URL.
        subprocess.CalledProcessError: If the git clone command fails.
    """
    if not repo_url.startswith("https://"):
        raise ValueError("Only HTTPS URLs are supported.")
    # Delete the existing folder if it exists
    
    if os.path.exists(repo_local_dir):
        print(f"Removing existing directory: {repo_local_dir}")
        shutil.rmtree(repo_local_dir, onerror=_on_rm_error)
        
    # Ensure the target folder exists
    os.makedirs(repo_local_dir, exist_ok=True)

    try:
        subprocess.run(["git", "clone", repo_url, repo_local_dir], check=True)
        print(f"Repository cloned into '{repo_local_dir}' successfully.")
    except subprocess.CalledProcessError as e:
        print(f"Failed to clone repository: {e}")
        raise
                
def _get_latest_git_tag_with_prefix(repo_path, prefix):
    try:
        # Change to the git repo directory
        original_dir = os.getcwd()
        os.chdir(repo_path)
    
        # Get all tags starting with the prefix, sorted in descending version order
        tags = subprocess.check_output(
            ["git", "tag", "--list", f"{prefix}*", "--sort=-v:refname"],
            text=True
        ).splitlines()
        return tags[0] if tags else None
    except subprocess.CalledProcessError:
        return None

def _next_progr_tagsuffix_from_git(repo_path, base_string, format_str="03d", format_str_prefix = "-"):
    latest_tag = _get_latest_git_tag_with_prefix(repo_path, base_string)

    if latest_tag:
        match = re.search(rf"{re.escape(base_string + format_str_prefix)}(\d+)$", latest_tag)
        last_number = int(match.group(1)) if match else 0
        print(f"LAST NNUMBER {last_number}")
    else:
        last_number = 0

    next_number = last_number + 1
    return format_str_prefix + format(next_number, format_str)
                            
        
def _is_tag_committed(tag_name, repo_path):
    """
    Check if a given Git tag exists in the repository.

    Args:
        tag_name (str): The name of the tag to check.
        repo_path (str): Path to the Git repository (default is current directory).

    Returns:
        bool: True if the tag exists (i.e., is committed), False otherwise.
    """
    try:
        result = subprocess.run(
            ['git', 'rev-parse', '--verify', f'refs/tags/{tag_name}'],
            cwd=repo_path,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )
        return result.returncode == 0
    except Exception as e:
        print(f"Error checking tag: {e}")
        return False        
 
def _list_git_changes(project_path):
    """
    Lists all tracked changes (modified, added, deleted) in a Git repository.
    
    :param project_path: Path to the Git project.
    :return: A dictionary containing lists of modified, added, and deleted files.
    """
    if not os.path.isdir(project_path):
        raise ValueError("Invalid project path")
    
    git_command = ["git", "status", "--porcelain"]
    try:
        result = subprocess.run(git_command, cwd=project_path, capture_output=True, text=True, check=True)
        changes = result.stdout.strip().split('\n')
        
        modified = []
        added = []
        deleted = []
        
        for change in changes:
            if change:
                status, file = change[:2].strip(), change[3:].strip()
                if status in ('M', 'MM'):
                    modified.append(file)
                elif status in ('A', 'AM'):
                    added.append(file)
                elif status in ('D', 'DM'):
                    deleted.append(file)
        
        return {"modified": modified, "added": added, "deleted": deleted}
    
    except subprocess.CalledProcessError as e:
        raise RuntimeError(f"Error executing git command: {e}")
        
def _is_tag_pushed(project_path, tag_name) -> bool:
    """
    Checks if a given tag is already pushed to the remote repository.
    
    :param project_path: Path to the local git repository.
    :param tag_name: Name of the tag to check.
    :return: True if the tag is pushed, False otherwise.
    """
    try:
        # Get the list of tags from the remote
        result = subprocess.run(
            ["git", "ls-remote", "--tags", "origin"],
            cwd=project_path,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=True
        )
        
        # Check if the tag exists in the remote
        return any(f"refs/tags/{tag_name}" in line for line in result.stdout.splitlines())
    except subprocess.CalledProcessError as e:
        print(f"Error checking remote tags: {e.stderr}")
        return False
        

def _is_last_commit_pushed(project_path):
    try:
        # Get the latest commit hash
        commit_hash = subprocess.check_output(
            ["git", "rev-parse", "HEAD"], cwd=project_path
        ).strip().decode("utf-8")

        # Check if the commit is in the remote
        remote_branches = subprocess.check_output(
            ["git", "branch", "-r", "--contains", commit_hash], cwd=project_path
        ).strip().decode("utf-8")

        # If the commit exists in remote branches, it's pushed
        return bool(remote_branches)
    except subprocess.CalledProcessError:
        return False              
                                        
def _resolve_placeholders(data, context=None):
    """ Recursively resolves placeholders in a dictionary using string formatting """
    if context is None:
        context = data  # Initial context is the entire data
    
    if isinstance(data, dict):
        return {key: _resolve_placeholders(value, context) for key, value in data.items()}
    elif isinstance(data, list):
        return [_resolve_placeholders(item, context) for item in data]
    elif isinstance(data, str):
        try:
            return data.format(**context)  # Use Python string formatting
        except KeyError:
            return data  # Return as-is if placeholders are unresolved
    else:
        return data  # Return non-string values unchanged
                        
def _execute_command(command, cwd):
    """Executes a shell command in a given directory."""
    try:
        subprocess.run(command, cwd=cwd, check=True, shell=True)
        logging.info(f"Executed: {' '.join(command)} in {cwd}")
        print(f"Executed: {' '.join(command)} in {cwd}")
    except subprocess.CalledProcessError as e:
        logging.error(f"Command failed: {e}")
        print(f"Command failed: {e}")
        raise
        
def _update_all_pom_properties(project, config):
    for root, _, files in os.walk(project["project_path"]):
        for file in files:
            if file == "pom.xml":
                file_path = os.path.join(root, file)
                _update_pom_property(file_path, project, config)

def _update_pom_property(file_path, project, config):
    # Parse XML while preserving comments
    parser = lxmlET.XMLParser(remove_comments=False)
    tree = lxmlET.parse(file_path, parser)
    root = tree.getroot()
    maven_namespace = config["maven_namespace"]
    
    # Define the XML namespace
    ns = {'maven': maven_namespace}
    ET.register_namespace('', ns['maven'])
    
    # Locate the <properties> section
    modified = False
    properties = root.find(".//maven:properties", ns)
    if properties is not None:     
        if project["type"] == "Maven":
            for configProperties in project["properties"]:                 
                property_element = properties.find(f"maven:{configProperties['property_name']}", ns)
                if property_element is not None:
                    property_element.text = configProperties['property_value']
                    logging.info(f"Updated {configProperties['property_name']} to {configProperties['property_value']}")
                    print(f"Updated {configProperties['property_name']} to {configProperties['property_value']}")
                    modified = True                
                else:
                    logging.info(f"Property {configProperties['property_name']} not found in pom.xml.")
                    print(f"Property {configProperties['property_name']} not found in pom.xml.")
    else:
        logging.info("No <properties> section found in pom.xml.")
        print("No <properties> section found in pom.xml.")
    if modified:
        tree.write(file_path, xml_declaration=True, encoding='utf-8')
        
def _update_maven_versions(project_path, dependencies, version, parent_version, maven_namespace):
    for root, _, files in os.walk(project_path):
        if 'pom.xml' in files:
            pom_path = os.path.join(root, 'pom.xml')
            # Parse XML while preserving comments
            parser = lxmlET.XMLParser(remove_comments=False)
            tree = lxmlET.parse(pom_path, parser)
            
            root_element = tree.getroot()
            
            # Define the XML namespaces
            # Define the XML namespace
            namespaces = {'m': maven_namespace}
            ET.register_namespace('', namespaces['m'])
            
            modified = False

            # Update the version of the project itself
            for project_version in root_element.findall("./m:version", namespaces):
                if project_version is not None:
                    project_version.text = version
                    modified = True
                    logging.info(f"Project version updated to {version}")
                    print(f"Project version updated to {version}")

            # Update the parent version
            for parent_tag in root_element.findall("./m:parent", namespaces):
                if parent_tag is not None:
                    for parent_vers in parent_tag.findall("./m:version", namespaces):                      
                        if parent_vers is not None:
                            parent_vers.text = parent_version
                            modified = True
                            logging.info(f"Parent version updated to {version}")
                            print(f"Parent version updated to {version}")
            
            
            for dependency in root_element.findall(".//m:dependency", namespaces):
                artifact_id_elem = dependency.find("m:artifactId", namespaces)
                version_elem = dependency.find("m:version", namespaces)
                
                if artifact_id_elem is not None and version_elem is not None:
                    for dependency in dependencies:
                        if artifact_id_elem.text in dependency["dependency_name"]:
                            version_elem.text = dependency["dependency_version"]
                            modified = True
                            logging.info(f"Dependency {dependency["dependency_name"]}  updated to {dependency["dependency_version"]}")
                            print(f"Dependency {dependency["dependency_name"]}  updated to {dependency["dependency_version"]}")            
            if modified:
                tree.write(pom_path, xml_declaration=True, encoding='utf-8')

def _update_maven_versions_from_yaml(project, config):    
    maven_namespace = config["maven_namespace"]
    project_path = project.get("project_path")
    parent_version = project.get("parent_version")
    dependencies = project.get("dependencies", [])
    version = project.get("version")
    if not project_path or not dependencies or not version or not maven_namespace:
        raise ValueError("YAML file must contain 'maven_namespace', 'project_path', 'parent_version', 'dependencies' and 'version' fields.")
    
    _update_maven_versions(project_path, dependencies, version, parent_version, maven_namespace)



def update_versions(project_filter = ''):
    """Reads the YAML file and processes each project."""
    with open("release_handler_config.yaml", "r") as file:
        config = yaml.safe_load(file)
    
    try:
        base_dir = config["base_dir"]
        for project in config["projects"]:
            if project_filter != '' and project_filter != project['name']:
                continue  
            if 'skip' in project and project['skip']:
                logging.info(f"Project {project['name']} is configured to be skipped")
                print(f"Project {project['name']} is configured to be skipped")
                continue
            repo_url = project["project_remote_git_url"]
            project_path = base_dir + '/' + project["name"]
            _clone_git_repo_delete_existent(repo_url, project_path)               
            if click.confirm(f"Update version for project {project['name']}?", default=True):
                changed = False
                if project["type"] == "Maven":
                    _update_all_pom_properties(project, config)
                    _update_maven_versions_from_yaml(project, config)
                    changed = True
                elif project["type"] == "Ant":
                    _update_ant_version(project["project_path"], project["version"], project["version_file"])
                    changed = True
                elif project["type"] == "Angular":
                    _update_angular_version(project["project_path"], project["version"], project["version_file"])
                    changed = True
                
                if changed:
                    changes =_list_git_changes(project["project_path"])
                    logging.info(f"Changes to commit {changes}")
                    print(f"Changes to commit {changes}")
                    if click.confirm(f"Commit changes for project {project['name']}?", default=False):
                        _execute_command(["git", "commit", "-am", f"Update project with version {project['version']}"] , project["project_path"])
                        logging.info(f"Updated project with version {project['version']}")
                        print(f"Updated project with version {project['version']}")
                else:
                    logging.info(f"No changes to commit {changes}")
                    print(f"No changes to commit {changes}")    
                    
    except Exception as e:
        logging.error(f"An error occurred: {e}")  
        print(f"An error occurred: {e}") 
        
        
def update_tags(project_filter = ''):
    """Tags each project with the appropriate tag name."""
    try:
        with open("release_handler_config.yaml", "r") as file:
            config = yaml.safe_load(file)
            resolved_config = _resolve_placeholders(config)
            
        base_dir = resolved_config["base_dir"]
        tag_progr_suffix = resolved_config["tag_progr_suffix"]
        tag_progr_suffix_format = resolved_config["tag_progr_suffix_format"]
        tag_progr_suffix_format_prefix = resolved_config["tag_progr_suffix_format_prefix"]       
        for project in resolved_config["projects"]:
            if project_filter != '' and project_filter != project['name']:
                continue  
            if 'skip' in project and project['skip']:
                logging.info(f"Project {project['name']} is configured to be skipped")
                print(f"Project {project['name']} is configured to be skipped")
                continue    
            tag = project["tag"]            
            if click.confirm(f"Create tag {tag} for project {project['name']}?", default=True):
                if _is_tag_committed(tag, project["project_path"]):
                    logging.info(f"Tag {tag} of project {project['name']} already commited")
                    print(f"Tag {tag} of project {project['name']} already commited")
                    continue
                repo_url = project["project_remote_git_url"]
                project_path = base_dir + '/' + project["name"]
                _clone_git_repo_delete_existent(repo_url, project_path)          
                
                if(tag_progr_suffix):                      
                    tag = tag + _next_progr_tagsuffix_from_git(project_path, tag, tag_progr_suffix_format, tag_progr_suffix_format_prefix) 
                print(f"Tagging {project['name']} with {tag}")                      
                _execute_command(["git", "tag", tag], project_path)
                logging.info(f"Tagged {project['name']} with {tag}")
                print(f"Tagged {project['name']} with {tag}")
                if _is_tag_pushed(project_path, tag):
                    logging.info(f"The {tag} for project {project['name']} is already pushed")
                    print(f"The {tag} for project {project['name']} is already pushed")
                    continue
                _execute_command(["git", "push", "origin", "tag", tag], project_path)
                logging.info(f"Pushed tag {tag} for project {project['name']}")
                print(f"Pushed tag {tag} for project {project['name']}")
    except Exception as ex:
        logging.error(f"An error occurred: {ex}")  
        print(f"An error occurred: {ex}")
        

def push_changes(project_filter=''):
    """Push committed changes to the remote repository for each project."""
    try:
        with open("release_handler_config.yaml", "r") as file:
            config = yaml.safe_load(file)
            resolved_config = _resolve_placeholders(config)

        base_dir = resolved_config["base_dir"]
        for project in resolved_config["projects"]:
            if project_filter != '' and project_filter != project['name']:
                continue
            if 'skip' in project and project['skip']:
                logging.info(f"Project {project['name']} is configured to be skipped")
                print(f"Project {project['name']} is configured to be skipped")
                continue

            project_path = os.path.join(base_dir, project["name"])
            if not _is_last_commit_pushed(project_path):
                if click.confirm(f"Push committed changes for {project['name']}?", default=True):
                    _execute_command(["git", "push"], project_path)
                    logging.info(f"Pushed committed changes for {project['name']}")
                    print(f"Pushed committed changes for {project['name']}")
                else:
                    print(f"Skipping push for {project['name']}")
            else:
                logging.info(f"No unpushed commits for {project['name']}")
                print(f"No unpushed commits for {project['name']}")
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        print(f"An error occurred: {e}")     

    
                 
if __name__ == "__main__":
    if len(sys.argv) > 1:
        if sys.argv[1] == "update_tags":
            if len(sys.argv) > 2 and sys.argv[2] != "":
                update_tags(sys.argv[2])
            else:
                update_tags()  
        if sys.argv[1] == "update_versions":
            if len(sys.argv) > 2 and sys.argv[2] != "":
                update_versions(sys.argv[2])
            else:
                update_versions()                 
        else:
             print("Wrong argument!")           
    else:
        raise ValueError("No arguments provided! Usage: release_handler.py <one of update_versions, create_tags, delete_tags, delete_tags, commit, remove_last_commit, reset, checkout_and_pull, compile_check>")